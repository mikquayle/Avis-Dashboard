import os
import json
import math
import requests
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
from openpyxl import Workbook, load_workbook

API_KEY = os.environ.get("GOOGLE_PLACES_API_KEY")
ANTHROPIC_KEY = os.environ.get("ANTHROPIC_API_KEY")

LOCATIONS = [
    {
        "name": "Avis Car Rental - Harry Reid Airport",
        "place_id": "ChIJL7BqEp3FyIARP_QZDyQ5HJI"
    },
]

DATA_FILE = "data/ratings.json"
MERGES_FILE = "data/merges.json"
LOG_FILE = "data/logs/Avis_Dashboard_Log.xlsx"
LAS_VEGAS = ZoneInfo("America/Los_Angeles")

SHOUTOUT_MIN_RATING = 4  # only 4 and 5 star reviews count toward employee shoutouts

# ---------- time helpers ----------

def get_lv_time():
    return datetime.now(LAS_VEGAS)

def get_work_day_key(lv_now):
    # Work day = 6:00 AM to 6:00 AM the next day
    if lv_now.hour < 6:
        return (lv_now - timedelta(days=1)).strftime("%Y-%m-%d")
    return lv_now.strftime("%Y-%m-%d")

def get_work_day_key_from_iso(iso_str):
    # Best-effort fallback, used only to migrate reviews counted before this
    # field existed. Newly counted reviews always use the work day the
    # scraper actually caught them on, not this.
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        lv = dt.astimezone(LAS_VEGAS)
    except Exception:
        lv = get_lv_time()
    return get_work_day_key(lv)

def get_week_key(dt):
    return dt.strftime("%Y-W%W")

def week_key_from_work_day(work_day_str):
    dt = datetime.strptime(work_day_str, "%Y-%m-%d")
    return get_week_key(dt)

# ---------- time-of-day review buckets ----------

BOUNDARY_HOURS = [6, 10, 14, 16, 20]

def bucket_counts_for_location(data, place_id, today_key, lv_hour, review_count):
    """
    Splits today's review count into 5 time-of-day windows (6-10, 10-14,
    14-16, 16-20, 20-6) by snapshotting the running review total each time
    the clock crosses one of the boundary hours. Each window's count is the
    difference between its start and end snapshot (or "so far" if the
    window hasn't closed yet).
    """
    baselines = data["daily_baselines"]
    snap = {}
    for bh in BOUNDARY_HOURS:
        key = place_id + "_tb" + str(bh) + "_" + today_key
        if key not in baselines and lv_hour >= bh:
            baselines[key] = review_count
        snap[bh] = baselines.get(key)

    def span(start_h, end_h):
        start_val = snap.get(start_h)
        if start_val is None:
            return 0
        if end_h is not None and snap.get(end_h) is not None:
            return max(0, snap[end_h] - start_val)
        return max(0, review_count - start_val)

    return {
        "6-10": span(6, 10),
        "10-14": span(10, 14),
        "14-16": span(14, 16),
        "16-20": span(16, 20),
        "20-6": span(20, None),
    }

# ---------- next rating-tier progress ----------

def rating_progress(data, place_id, rating, review_count):
    """
    Estimates how many 5-star reviews are needed to push the rounded rating
    up to the next tenth (e.g. 4.0 -> 4.1), and tracks progress toward that
    since the current tier was first reached. Progress resets automatically
    the moment the rounded rating advances to a new tier.
    """
    if rating is None or review_count is None:
        return None

    tier = round(rating, 1)
    next_goal = round(tier + 0.1, 1)

    if next_goal >= 5.0:
        needed_now = 0
    else:
        needed_now = max(0, math.ceil(review_count * (next_goal - rating) / (5 - next_goal)))

    tracking = data.setdefault("rating_goal_tracking", {}).setdefault(place_id, {})
    if tracking.get("tier") != tier:
        tracking["tier"] = tier
        tracking["next_goal"] = next_goal
        tracking["baseline_needed"] = needed_now if needed_now > 0 else 1

    baseline_needed = tracking.get("baseline_needed") or (needed_now if needed_now > 0 else 1)
    progress_pct = 0
    if baseline_needed > 0:
        progress_pct = max(0, min(100, round((1 - (needed_now / baseline_needed)) * 100)))

    return {
        "current_tier": tier,
        "next_goal": next_goal,
        "five_star_reviews_needed": needed_now,
        "progress_pct": progress_pct,
    }

def rebuild_mentions(data, place_id, merges, today_key):
    """
    Recomputes employee shoutout totals from scratch every run, based on the
    work day each review was actually counted on (not the day the reviewer
    originally posted it on Google). Only 4+ star reviews count.
    Returns (today's week key, today's mention counts).
    """
    reviews = data.get("reviews", {}).get(place_id, [])
    weekly_history = {}
    today_counts = {}

    for r in reviews:
        work_day = r.get("counted_work_day")
        rating = r.get("rating", 0)
        names = r.get("employee_names") or []
        if not work_day or not names or rating < SHOUTOUT_MIN_RATING:
            continue

        wk = week_key_from_work_day(work_day)
        seen = set()
        for raw in names:
            if not raw or not raw.strip():
                continue
            emp_key, display = canonicalize(raw, merges)
            if emp_key in seen:
                continue
            seen.add(emp_key)

            bucket = weekly_history.setdefault(wk, {})
            entry = bucket.setdefault(emp_key, {"display_name": display, "count": 0})
            entry["count"] += 1

            if work_day == today_key:
                t_entry = today_counts.setdefault(
                    emp_key, {"display_name": display, "count": 0, "last_mentioned": today_key}
                )
                t_entry["count"] += 1

    data.setdefault("weekly_mentions_history", {})[place_id] = weekly_history
    data.setdefault("employee_mentions", {})[place_id] = today_counts
    return week_key_from_work_day(today_key), today_counts

# ---------- Google Places ----------

def fetch_place(place_id):
    url = "https://places.googleapis.com/v1/places/" + place_id
    headers = {
        "X-Goog-Api-Key": API_KEY,
        # IMPORTANT: list review sub-fields explicitly. Requesting the bare
        # "reviews" field can return incomplete/empty review objects.
        "X-Goog-FieldMask": "id,displayName,rating,userRatingCount,formattedAddress,"
                             "reviews.text,reviews.rating,reviews.publishTime,reviews.authorAttribution",
    }
    response = requests.get(url, headers=headers, timeout=15)
    response.raise_for_status()
    data = response.json()
    print("Reviews fetched: " + str(len(data.get("reviews", []))))
    return data

def extract_employee_names(review_text):
    if not ANTHROPIC_KEY:
        return []
    if not review_text or len(review_text.strip()) < 5:
        return []
    try:
        prompt = (
            "Read this car rental review and extract any employee first names or nicknames mentioned. "
            "Return ONLY a JSON array of name strings, nothing else. "
            "If no employee names are mentioned, return an empty array []. "
            "Examples: [\"Mike\", \"Sandra\"] or [\"Big John\"] or []\n\n"
            "Review: " + review_text
        )
        response = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "Content-Type": "application/json",
                "x-api-key": ANTHROPIC_KEY,
                "anthropic-version": "2023-06-01",
            },
            json={
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 100,
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=15,
        )
        response.raise_for_status()
        text = response.json()["content"][0]["text"].strip()
        text = text.replace("```json", "").replace("```", "").strip()
        names = json.loads(text)
        if isinstance(names, list):
            return [str(n).strip() for n in names if n]
        return []
    except Exception as e:
        print("Name extraction error: " + str(e))
        return []

# ---------- persistence ----------

def load_existing_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    return {}

def save_data(data):
    os.makedirs("data", exist_ok=True)
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2)

def load_merges():
    if os.path.exists(MERGES_FILE):
        with open(MERGES_FILE, "r") as f:
            return json.load(f)
    return {}

def canonicalize(name, merges):
    key = name.lower().strip()
    canon_key = merges.get(key, key)
    display = merges.get("__display__" + canon_key, name.strip())
    return canon_key, display

# ---------- Excel logging ----------

def init_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Settings"
    ws["A1"] = "Avis Dashboard — Config"
    ws["A3"] = "Daily Review Goal"
    ws["B3"] = 50
    ws["A4"] = "Gift Card Shoutout Threshold (mentions/day)"
    ws["B4"] = 3
    ws["A5"] = "Work Day Window"
    ws["B5"] = "6:00 AM – 6:00 AM (next day)"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20

    ds = wb.create_sheet("Daily Summary")
    ds.append(["Date", "Rating (End of Day)", "Total Reviews (End of Day)",
               "True New Reviews Today", "Reviews w/ Text Captured", "Capture Rate", "Daily Goal Met?"])

    sd = wb.create_sheet("Shoutouts Daily")
    sd.append(["Date", "Week", "Employee (merged name)", "Mentions Today", "Threshold", "Gift Card Earned?"])

    sw = wb.create_sheet("Shoutouts Weekly")
    sw.append(["Week", "Employee (merged name)", "Total Mentions", "Threshold", "Gift Cards Earned (days met)"])
    return wb

def log_day_close(closed_date, week_key, rating, review_count_end, true_new, captured_count, employee_day_counts):
    os.makedirs("data/logs", exist_ok=True)
    wb = load_workbook(LOG_FILE) if os.path.exists(LOG_FILE) else init_workbook()

    ds = wb["Daily Summary"]
    r = ds.max_row + 1
    ds.cell(r, 1, closed_date)
    ds.cell(r, 2, rating)
    ds.cell(r, 3, review_count_end)
    ds.cell(r, 4, true_new)
    ds.cell(r, 5, captured_count)
    ds.cell(r, 6, "=IFERROR(E{0}/D{0},0)".format(r)).number_format = "0%"
    ds.cell(r, 7, '=IF(D{0}>=Settings!$B$3,"YES","NO")'.format(r))

    sd = wb["Shoutouts Daily"]
    for emp_key, info in employee_day_counts.items():
        r2 = sd.max_row + 1
        sd.cell(r2, 1, closed_date)
        sd.cell(r2, 2, week_key)
        sd.cell(r2, 3, info["display_name"])
        sd.cell(r2, 4, info["count"])
        sd.cell(r2, 5, "=Settings!$B$4")
        sd.cell(r2, 6, '=IF(D{0}>=E{0},"YES","NO")'.format(r2))

    sw = wb["Shoutouts Weekly"]
    existing = {(sw.cell(row=r, column=1).value, sw.cell(row=r, column=2).value)
                for r in range(2, sw.max_row + 1)}
    for emp_key, info in employee_day_counts.items():
        key = (week_key, info["display_name"])
        if key not in existing:
            r3 = sw.max_row + 1
            sw.cell(r3, 1, week_key)
            sw.cell(r3, 2, info["display_name"])
            sw.cell(r3, 3, "=SUMIFS('Shoutouts Daily'!D:D,'Shoutouts Daily'!B:B,A{0},'Shoutouts Daily'!C:C,B{0})".format(r3))
            sw.cell(r3, 4, "=Settings!$B$4")
            sw.cell(r3, 5, '=COUNTIFS(\'Shoutouts Daily\'!B:B,A{0},\'Shoutouts Daily\'!C:C,B{0},\'Shoutouts Daily\'!D:D,">="&D{0})'.format(r3))
            existing.add(key)

    wb.save(LOG_FILE)
    print("Excel log updated for " + closed_date)

# ---------- daily star (existing weekly star feature, now on 6am cutover) ----------

def award_daily_star(data, place_id, closed_date, week_key):
    award_key = place_id + "_star_" + closed_date
    if data.get("daily_stars", {}).get(award_key):
        return
    employees = data.get("employee_mentions", {}).get(place_id, {})
    if not employees:
        return
    top = max(employees.items(), key=lambda x: x[1].get("count", 0), default=None)
    if not top or top[1].get("count", 0) == 0:
        return
    top_key = top[0]
    data.setdefault("weekly_stars", {}).setdefault(place_id, {}).setdefault(week_key, {})
    week_data = data["weekly_stars"][place_id][week_key]
    if sum(week_data.values()) >= 7:
        data["daily_stars"][award_key] = "max_week"
        return
    if week_data.get(top_key, 0) >= 3:
        data["daily_stars"][award_key] = "max_emp"
        return
    week_data[top_key] = week_data.get(top_key, 0) + 1
    data["daily_stars"][award_key] = top_key
    print("Star awarded to: " + top[1]["display_name"])

# ---------- cutover (close out the workday that just ended) ----------

def do_cutover(data, place_id, lv_now, review_count_now, rating_now):
    if lv_now.hour != 6:
        return
    closed_date = (lv_now - timedelta(days=1)).strftime("%Y-%m-%d")
    cutover_key = place_id + "_cutover_" + closed_date
    data.setdefault("cutover_done", {})
    if data["cutover_done"].get(cutover_key):
        return

    closed_baseline_key = place_id + "_" + closed_date
    closed_baseline = data.get("daily_baselines", {}).get(closed_baseline_key, review_count_now)
    true_new = max(0, review_count_now - closed_baseline)

    captured_count = data.get("captured_counts", {}).get(place_id, {}).get(closed_date, 0)

    employees = data.get("employee_mentions", {}).get(place_id, {})
    employee_day_counts = {k: {"display_name": v["display_name"], "count": v["count"]}
                            for k, v in employees.items() if v.get("count", 0) > 0}

    week_key = get_week_key(lv_now - timedelta(days=1))

    award_daily_star(data, place_id, closed_date, week_key)
    log_day_close(closed_date, week_key, rating_now, review_count_now, true_new, captured_count, employee_day_counts)

    # Employee mention counts are rebuilt from scratch every run (see
    # rebuild_mentions), so no manual reset is needed here.
    data.setdefault("captured_counts", {}).setdefault(place_id, {})[closed_date] = 0

    data["cutover_done"][cutover_key] = True
    print("Cutover complete for " + closed_date + " — true new reviews: " + str(true_new))

# ---------- main ----------

def main():
    if not API_KEY:
        raise EnvironmentError("GOOGLE_PLACES_API_KEY secret is not set.")

    lv_now = get_lv_time()
    today_key = get_work_day_key(lv_now)
    today_date = lv_now.strftime("%Y-%m-%d")
    now_str = lv_now.strftime("%Y-%m-%d %H:%M")
    lv_hour = lv_now.hour

    data = load_existing_data()
    for k in ["locations", "history", "reviews", "daily_baselines", "employee_mentions",
              "counted_review_ids", "daily_stars", "weekly_stars", "captured_counts",
              "cutover_done", "recent_reviews"]:
        if k not in data:
            data[k] = {}

    merges = load_merges()

    for location in LOCATIONS:
        name = location["name"]
        place_id = location["place_id"]
        print("Fetching: " + name)

        place = fetch_place(place_id)
        rating = place.get("rating")
        review_count = place.get("userRatingCount")
        address = place.get("formattedAddress", "")
        raw_reviews = place.get("reviews", [])

        print("Rating: " + str(rating) + " (" + str(review_count) + ")")

        do_cutover(data, place_id, lv_now, review_count, rating)

        baseline_key = place_id + "_" + today_key
        if baseline_key not in data["daily_baselines"] and lv_hour >= 6:
            data["daily_baselines"][baseline_key] = review_count
            print("Set daily baseline: " + str(review_count))

        baseline = data["daily_baselines"].get(baseline_key, review_count)
        reviews_today = max(0, review_count - baseline)

        review_buckets = bucket_counts_for_location(data, place_id, today_key, lv_hour, review_count)
        print("Today: " + str(reviews_today) + " | Buckets: " + json.dumps(review_buckets))

        progress = rating_progress(data, place_id, rating, review_count)
        if progress:
            print("Next goal: " + str(progress["next_goal"]) +
                  " (" + str(progress["five_star_reviews_needed"]) + " 5-star reviews needed, " +
                  str(progress["progress_pct"]) + "% there)")

        data["locations"][place_id] = {
            "name": name, "address": address, "rating": rating,
            "review_count": review_count, "reviews_today": reviews_today,
            "review_buckets": review_buckets, "rating_progress": progress,
            "last_updated": now_str, "today_key": today_key,
        }

        if place_id not in data["history"]:
            data["history"][place_id] = []
        existing_dates = [h["date"] for h in data["history"][place_id]]
        if today_date not in existing_dates:
            data["history"][place_id].append({"date": today_date, "rating": rating, "review_count": review_count})
        else:
            for h in data["history"][place_id]:
                if h["date"] == today_date:
                    h["rating"] = rating
                    h["review_count"] = review_count

        data["reviews"].setdefault(place_id, [])
        data["employee_mentions"].setdefault(place_id, {})
        data["counted_review_ids"].setdefault(place_id, [])
        data["captured_counts"].setdefault(place_id, {})

        counted_ids = set(data["counted_review_ids"][place_id])
        stored_by_id = {r["id"]: r for r in data["reviews"][place_id]}

        fresh_reviews = []
        for review in raw_reviews:
            author = review.get("authorAttribution", {}).get("displayName", "Anonymous")
            publish_time = review.get("publishTime", "")
            review_id = author + "_" + publish_time
            text_field = review.get("text", "")
            text = text_field.get("text", "") if isinstance(text_field, dict) else str(text_field)
            star_rating = review.get("rating", 0)

            if review_id in counted_ids:
                prev = stored_by_id.get(review_id, {})
                names = prev.get("employee_names", [])
                # Older reviews counted before this field existed fall back
                # to an estimate from their original post time.
                counted_work_day = prev.get("counted_work_day") or get_work_day_key_from_iso(publish_time)
                counted_at = prev.get("counted_at")
                review_date = prev.get("date", today_date)
                print("Already counted: " + author)
            else:
                print("NEW: " + author + " " + str(star_rating) + "★ (" + publish_time[:10] + ")")
                names = extract_employee_names(text)
                counted_work_day = today_key
                counted_at = lv_now.isoformat()
                review_date = today_date

                data["captured_counts"][place_id][today_key] = \
                    data["captured_counts"][place_id].get(today_key, 0) + 1

                if names:
                    print("  Employees: " + str(names))
                counted_ids.add(review_id)

            fresh_reviews.append({
                "id": review_id, "author": author, "rating": star_rating,
                "text": text, "publish_time": publish_time,
                "date": review_date, "employee_names": names if names else [],
                "counted_work_day": counted_work_day, "counted_at": counted_at,
            })

        fresh_reviews.sort(key=lambda r: r.get("publish_time", ""), reverse=True)
        fresh_ids = {r["id"] for r in fresh_reviews}
        older = [r for r in data["reviews"][place_id] if r["id"] not in fresh_ids]
        data["reviews"][place_id] = fresh_reviews + older
        data["reviews"][place_id] = data["reviews"][place_id][:100]
        data["counted_review_ids"][place_id] = list(counted_ids)

        data["recent_reviews"][place_id] = [
            {
                "author": r["author"],
                "rating": r["rating"],
                "text": r["text"],
                "publish_time": r["publish_time"],
                "date": r["date"],
                "employee_names": r.get("employee_names", []),
                "counted_at": r.get("counted_at"),
            }
            for r in data["reviews"][place_id][:10]
        ]

        today_week, current = rebuild_mentions(data, place_id, merges, today_key)
        print("Shoutouts today (4+ star, " + today_key + "): " +
              (", ".join(sorted(current.keys())) if current else "(none yet today)"))

        print("Total reviews stored: " + str(len(data["reviews"][place_id])))
        print("Total counted IDs: " + str(len(counted_ids)))

    save_data(data)
    print("Saved.")

if __name__ == "__main__":
    main()
